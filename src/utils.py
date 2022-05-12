# Funciones de utilidad: lectura de datos, representación básica, etc.

import openpyxl


def read_file(filename, size):
    """
    Función que permite la lectura de un fichero .csv o .txt.

    :param filename: Nombre del fichero que contiene los datos.
    :param size: Número de datos a leer.

    :return: data: Estructura de datos del problema.
    :return: max_capacity: Capacidad máxima de los vehículos.
    :return: max_time: Jornada laboral máxima (en minutos).
    :return: max_vehicles: Número máximo de vehículos en la solución admisible.
    :return: depots: Índices donde se encuentran los distintos depósitos del problema (en este caso solo 1).
    """

    # Inicialización de estructura de datos
    data = {'customer': [], 'x_coord': [], 'y_coord': [], 'demand': [], 'ready_time': [],
            'due_date': [], 'service_time': []}

    # Lectura de fichero
    with open(filename, 'r') as file:

        # Recorrido del fichero
        for i, line in enumerate(file):

            # Número de vehículos y capacidad máxima de los mismos
            if i == 4:
                max_vehicles, max_capacity = map(int, line.split())

            # Depósito
            elif i == 9:
                line_numbers = list(map(int, line.split()))
                data['customer'].append(line_numbers[0])
                data['x_coord'].append(line_numbers[1])
                data['y_coord'].append(line_numbers[2])
                data['demand'].append(line_numbers[3])
                data['ready_time'].append(line_numbers[4])
                data['due_date'].append(line_numbers[5])
                data['service_time'].append(line_numbers[6])

                max_time = float(line_numbers[5])
                depots = [line_numbers[0]]

            # Datos de paradas
            elif (i >= 10) and (i < 10 + size):
                line_numbers = list(map(int, line.split()))
                data['customer'].append(line_numbers[0])
                data['x_coord'].append(line_numbers[1])
                data['y_coord'].append(line_numbers[2])
                data['demand'].append(line_numbers[3])
                data['ready_time'].append(line_numbers[4])
                data['due_date'].append(line_numbers[5])
                data['service_time'].append(line_numbers[6])

            # (Máximo tamaño alcanzado → Salir del bucle)
            elif i >= 10 + size:
                break
    file.close()

    return data, max_capacity, max_time, max_vehicles, depots


def read_file_excel(filename, size):
    """
    Función que permite la lectura de un fichero .xlsx.

    :param filename: Nombre del fichero que contiene los datos.
    :param size: Número de datos a leer.

    :return: data: Estructura de datos del problema.
    :return: max_capacity: Capacidad máxima de los vehículos.
    :return: max_time: Jornada laboral máxima (en minutos).
    :return: max_vehicles: Número máximo de vehículos en la solución admisible.
    :return: depots: Índices donde se encuentran los distintos depósitos del problema (en este caso solo 1).
    """

    # Inicialización de estructura de datos
    data = {'customer': [], 'x_coord': [], 'y_coord': [], 'demand': [], 'ready_time': [],
            'due_date': [], 'service_time': []}

    # Lectura de hoja
    file = openpyxl.load_workbook(filename)
    sheet = file['Hoja1']

    # Número de vehículos y capacidad máxima de los mismos
    max_vehicles = sheet.cell(4, 4).value
    max_capacity = sheet.cell(6, 4).value

    # Depósito
    data['customer'].append(int(sheet.cell(11, 2).value))
    data['x_coord'].append(int(sheet.cell(11, 3).value))
    data['y_coord'].append(int(sheet.cell(11, 4).value))
    data['demand'].append(int(sheet.cell(11, 5).value))
    data['ready_time'].append(int(sheet.cell(11, 6).value))
    data['due_date'].append(int(sheet.cell(11, 7).value))
    data['service_time'].append(int(sheet.cell(11, 8).value))

    max_time = float(data['due_date'][-1])
    depots = [data['customer'][-1]]

    # Datos de paradas
    for i in range(12, min(12 + size, sheet.max_row + 1)):
        data['customer'].append(int(sheet.cell(i, 2).value))
        data['x_coord'].append(int(sheet.cell(i, 3).value))
        data['y_coord'].append(int(sheet.cell(i, 4).value))
        data['demand'].append(int(sheet.cell(i, 5).value))
        data['ready_time'].append(int(sheet.cell(i, 6).value))
        data['due_date'].append(int(sheet.cell(i, 7).value))
        data['service_time'].append(int(sheet.cell(i, 8).value))

    file.close()

    return data, max_capacity, max_time, max_vehicles, depots
